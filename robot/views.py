import mimetypes
import re
from datetime import date, datetime
from io import BytesIO
import io
import openpyxl
from django.core.files.base import ContentFile
from django.core.mail import send_mail
from django.db import transaction
from django.template.loader import render_to_string
from openpyxl.drawing.image import Image as OpenPyXLImage
import os
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import JsonResponse, HttpResponse, FileResponse, HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, get_user_model
from django.utils.timezone import make_naive, localtime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from django.db.models import Q
from openpyxl.workbook import Workbook

from .forms import UserRegistrationForm, ChangePasswordForm, EditProfileForm, EditUserProfileForm, CompetitionForm, \
    KitForm, KitImageForm, SponsorForm, FeedbackForm, GuideFileForm, RegistrationForm, EditRegistrationForm

from django.contrib.auth.forms import AuthenticationForm, PasswordResetForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.utils import timezone

from .models import UserProfile, Competition, Registration, Notification, CustomUser, CompetitionResult, Kit, Sponsor, \
    Feedback, GuideFile, Team
from .utils import generate_random_code


def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()  # Lưu người dùng mới

            # Phân biệt user và admin
            if not user.is_superuser:
                login(request, user)  # Tự động đăng nhập cho người dùng thường

                # Tạo thông báo cho người dùng hiện tại (không phải admin)
                Notification.objects.create(
                    user=user,
                    title="Chào mừng đến với ứng dụng quản lý cuộc thi robot",
                    content="Chào mừng bạn đã đến với ứng dụng thông tin quản lý các cuộc thi robot. Hãy bắt đầu khám phá các tính năng!",
                    created_at=timezone.now(),
                    notification_type='user'
                )

            # Tạo thông báo cho admin về người dùng mới
            admin_users = get_user_model().objects.filter(is_superuser=True)  # Lấy danh sách admin
            for admin in admin_users:
                Notification.objects.create(
                    user=admin,
                    title="Thông báo người dùng đăng ký mới",
                    content=f"Người dùng mới {user.full_name} với username {user.username} đã đăng ký thành công tài khoản trên hệ thống.",
                    created_at=timezone.now(),
                    notification_type='admin'
                )

            return redirect('home')  # Thay 'home' bằng tên URL mà bạn muốn chuyển hướng sau khi đăng ký
    else:
        form = UserRegistrationForm()

    return render(request, 'account/register.html', {'form': form})


def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')  # Chuyển hướng đến trang sau khi đăng nhập thành công
    else:
        form = AuthenticationForm()
    return render(request, 'account/login.html', {'form': form})


@login_required
def home(request):
    return render(request, 'home.html')


def user_logout(request):
    logout(request)
    return redirect('login')  # Chuyển hướng đến trang đăng nhập sau khi đăng xuất


@login_required
def notification_view(request):
    # Lấy danh sách thông báo của người dùng và sắp xếp theo thời gian tạo
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    notifications_unread = notifications.filter(is_read=False).exists()

    # Tạo phân trang với mỗi trang hiển thị 7 thông báo
    paginator = Paginator(notifications, 7)  # Hiển thị 7 thông báo mỗi trang
    page_number = request.GET.get('page')  # Lấy số trang hiện tại từ URL
    page_obj = paginator.get_page(page_number)  # Lấy các thông báo tương ứng với trang hiện tại

    return render(request, 'notification.html', {
        'page_obj': page_obj,  # Thay đổi từ 'notifications' sang 'page_obj'
        'notifications_unread': notifications_unread
    })


@login_required
def mark_as_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()

    # Trả về phản hồi JSON để AJAX có thể xử lý
    return JsonResponse({'status': 'success'})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = ChangePasswordForm(request.POST, user=request.user)
        if form.is_valid():
            current_password = form.cleaned_data['current_password']
            new_password = form.cleaned_data['new_password']

            # Kiểm tra mật khẩu cũ và cập nhật mật khẩu mới
            user = request.user
            if user.check_password(current_password):
                user.set_password(new_password)
                user.save()
                update_session_auth_hash(request, user)  # Cập nhật session để tránh đăng xuất

                # Change password
                if user.is_superuser:
                    Notification.objects.create(
                        user=user,
                        title="Cập nhật mật khẩu thành công",
                        content="Bạn đã cập nhật thành công mật khẩu mới cho tài khoản của bạn.",
                        notification_type='admin',
                        created_at=timezone.now()
                    )
                else:
                    Notification.objects.create(
                        user=user,
                        title="Cập nhật mật khẩu thành công",
                        content="Bạn đã cập nhật thành công mật khẩu mới cho tài khoản của bạn.",
                        notification_type='user',
                        created_at=timezone.now()
                    )

                # messages.success(request, 'Mật khẩu đã được thay đổi thành công.')
                return redirect('home')
            else:
                form.add_error('current_password', 'Mật khẩu cũ không đúng.')
        else:
            print(form.errors)
            messages.error(request, 'Vui lòng kiểm tra các lỗi trong biểu mẫu.')
    else:
        form = ChangePasswordForm(user=request.user)

    return render(request, 'account/change_password.html', {'form': form})


@login_required
def user_profile(request):
    user = request.user
    # Kiểm tra và tạo UserProfile nếu chưa tồn tại
    profile, created = UserProfile.objects.get_or_create(user=user)

    context = {
        'profile': profile,
    }
    return render(request, 'profile/user_profile.html', context)


@login_required
def edit_profile(request):
    user = request.user
    profile, created = UserProfile.objects.get_or_create(user=user)

    if request.method == 'POST':
        user_form = EditProfileForm(request.POST, request.FILES, instance=user)
        profile_form = EditUserProfileForm(request.POST, request.FILES, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()

            # Tạo thông báo cho người dùng hoặc admin dựa trên vai trò
            if user.is_superuser:
                Notification.objects.create(
                    user=user,
                    title="Cập nhật thông tin cá nhân thành công",
                    content="Thông tin cá nhân của admin đã được cập nhật thành công.",
                    created_at=timezone.now(),
                    notification_type='admin'
                )
            else:
                Notification.objects.create(
                    user=user,
                    title="Cập nhật thông tin cá nhân thành công",
                    content="Thông tin cá nhân của bạn đã được cập nhật thành công.",
                    created_at=timezone.now(),
                    notification_type='user'
                )

            messages.success(request, 'Thông tin cá nhân đã được cập nhật.')
            return redirect('user_profile')  # Điều hướng về trang hồ sơ cá nhân sau khi chỉnh sửa
        else:
            print(user_form.errors)
            print(profile_form.errors)

    else:
        user_form = EditProfileForm(instance=user)
        profile_form = EditUserProfileForm(instance=profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
    }
    return render(request, 'profile/edit_profile.html', context)


@login_required
def contest_list(request):
    # competitions = Competition.objects.filter(is_active=True)
    competitions = Competition.objects.filter()
    today = date.today()
    return render(request, 'competition/contest.html', {'competitions': competitions, 'today': today})


@login_required
def competition_detail(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)
    guide_files = competition.guide_files.filter(is_confirmed=True).order_by('-uploaded_at')
    today = timezone.now().date()

    # Tìm đội mà người dùng tham gia trong cuộc thi này
    team = Team.objects.filter(members=request.user, competition=competition).first()

    # Nếu có đội, tìm đăng ký của đội đó
    if team:
        registration = Registration.objects.filter(team=team, competition=competition).first()
    else:
        registration = None

    # Đếm số lượng đội đã đăng ký nhưng chưa bị hủy trong cuộc thi này
    current_participants = Registration.objects.filter(
        competition=competition,
        is_cancelled=False
    ).count()

    return render(request, 'competition/competition_detail.html', {
        'competition': competition,
        'today': today,
        'registration': registration,
        'current_participants': current_participants,
        'guide_files': guide_files,
    })


@login_required
def register_competition(request, competition_id):
    # Lấy thông tin cuộc thi theo ID đã cho
    competition = get_object_or_404(Competition, id=competition_id)
    today = timezone.now().date()  # Lấy ngày hiện tại
    user = request.user

    # Kiểm tra khoảng thời gian đăng ký
    if not (competition.registration_start_date <= today <= competition.registration_end_date):
        messages.error(request, 'Ngoài khoảng thời gian đăng ký.')  # Thông báo nếu không trong khoảng thời gian
        return redirect('competition_detail', competition_id=competition.id)  # Quay lại trang chi tiết cuộc thi

    # Kiểm tra số lượng đội đăng ký
    current_participants = competition.teams.count()  # Đếm số đội hiện tại đã đăng ký
    if competition.max_participants is not None and current_participants >= competition.max_participants:
        messages.error(request, 'Số lượng đội đã đầy.')  # Thông báo nếu số lượng đội đã đầy
        return redirect('competition_detail', competition_id=competition.id)  # Quay lại trang chi tiết cuộc thi

    if request.method == 'POST':  # Kiểm tra xem có phải là request POST không
        form = RegistrationForm(request.POST, request.FILES)  # Khởi tạo form với dữ liệu từ request

        if form.is_valid():  # Kiểm tra tính hợp lệ của form
            try:
                with transaction.atomic():
                    # Tạo đối tượng Team mới
                    print(form.cleaned_data['team_name'])
                    team = Team.objects.create(
                        name=form.cleaned_data['team_name'],  # Lấy tên đội từ form
                        competition=competition,  # Liên kết với cuộc thi
                        coach=form.cleaned_data['coach_name']  # Lấy thông tin huấn luyện viên từ form
                    )

                    # Khởi tạo danh sách để lưu trữ thông tin
                    student_names = []
                    relationships = []
                    guardian_names = []
                    guardian_phones = []
                    guardian_emails = []
                    student_phones = []
                    student_emails = []
                    student_classes = []
                    birth_years = []
                    genders = []
                    ethnicities = []
                    addresses = []
                    student_photos = []

                    # Lấy dữ liệu từ POST cho số lượng thành viên
                    member_count = int(request.POST['member_count'])
                    print("member_count", member_count)
                    # Lấy thông tin từ POST
                    city = request.POST.get('city')
                    coach_name = request.POST.get('coach_name')
                    coach_phone = request.POST.get('coach_phone')
                    coach_unit = request.POST.get('coach_unit')
                    competition_group = request.POST.get('competition_group')
                    region = request.POST.get('region')
                    team_email = request.POST.get('team_email')
                    team_name = request.POST.get('team_name')

                    for i in range(1, member_count + 1):
                        # Lấy thông tin cho từng học sinh
                        student_names.append(request.POST.get(f'student_name[{i}]'))
                        relationships.append(request.POST.get(f'relationship[{i}]'))
                        guardian_names.append(request.POST.get(f'guardian_name[{i}]'))
                        guardian_phones.append(request.POST.get(f'guardian_phone[{i}]'))
                        guardian_emails.append(request.POST.get(f'guardian_email[{i}]'))
                        student_phones.append(request.POST.get(f'student_phone[{i}]'))
                        student_emails.append(request.POST.get(f'student_email[{i}]'))
                        student_classes.append(request.POST.get(f'student_class[{i}]'))
                        birth_years.append(request.POST.get(f'birth_year[{i}]'))
                        genders.append(request.POST.get(f'gender[{i}]'))
                        ethnicities.append(request.POST.get(f'ethnicity[{i}]'))
                        addresses.append(request.POST.get(f'address[{i}]'))
                        student_photos.append(request.FILES.get(f'student_photo[{i}]'))

                    # Lưu danh sách tên vào trường members của Team, phân tách bằng dấu phẩy
                    team.members = ', '.join(student_names)  # Cập nhật danh sách thành viên của đội
                    team.save()  # Lưu lại đối tượng Team

                    # Lặp qua từng thành viên và tạo Registration cho từng thành viên
                    for i in range(len(student_names)):
                        # Tạo đối tượng Registration và lưu thông tin
                        registration = Registration.objects.create(
                            competition=competition,
                            team=team,
                            user=user,
                            guardian_name=guardian_names[i],
                            relationship=relationships[i],
                            guardian_phone=guardian_phones[i],
                            guardian_email=guardian_emails[i],
                            student_name=student_names[i],
                            student_phone=student_phones[i],
                            student_email=student_emails[i],
                            student_class=student_classes[i],
                            birth_year=birth_years[i],
                            gender=genders[i],
                            ethnicity=ethnicities[i],
                            address=addresses[i],
                            student_photo=student_photos[i],
                            city=city,
                            coach_name=coach_name,
                            coach_phone=coach_phone,
                            coach_unit=coach_unit,
                            competition_group=competition_group,
                            region=region,
                            team_email=team_email,
                            team_name=team_name,
                            member_count=member_count,
                            registration_date=timezone.now(),
                            status="Đăng ký thành công"
                        )

                    # Thông báo thành công
                    messages.success(request, f'Đăng ký đội {team.name} thành công!')  # Thông báo cho người dùng

                    # Tạo thông báo cho người dùng
                    Notification.objects.create(
                        user=request.user,
                        title="Đăng ký đội thành công",
                        content=f"Bạn đã đăng ký đội {team.name} cho cuộc thi {competition.name}.",
                        created_at=timezone.now(),
                        notification_type='user'
                    )

                    # Thông báo cho admin
                    admin_users = get_user_model().objects.filter(is_superuser=True)  # Lấy danh sách admin

                    # Tạo danh sách các thông báo dành cho admin
                    admin_notifications = [
                        Notification(
                            user=admin,
                            title="Thông báo đăng ký đội mới",
                            content=f"Đội {team.name} đã đăng ký cuộc thi {competition.name}.",
                            created_at=timezone.now(),
                            notification_type='admin'
                        )
                        for admin in admin_users
                    ]

                    # Sử dụng bulk_create để tạo tất cả thông báo cùng lúc
                    Notification.objects.bulk_create(admin_notifications)

                    # Redirect đến trang chi tiết cuộc thi
                    return redirect('competition_detail', competition_id=competition.id)

            except Exception as e:
                messages.error(request, f"Có lỗi xảy ra: {str(e)}")  # Hiển thị lỗi nếu xảy ra ngoại lệ
                return redirect('competition_detail', competition_id=competition.id)
        else:
            # In ra các lỗi trong form nếu form không hợp lệ
            print(form.errors)  # In lỗi ra console
            participants_data = request.POST.getlist('student_name')  # Lưu dữ liệu người dùng nhập vào
            guardians_data = request.POST.getlist('guardian_name')  # Lưu dữ liệu người dùng nhập vào

            # Truyền lại dữ liệu người dùng nhập vào cho các trường trong context
            return render(request, 'competition/register_team.html', {
                'competition': competition,
                'form': form,
                'today': today,
                'participants_data': participants_data,
                'guardians_data': guardians_data,
            })

    else:
        form = RegistrationForm()  # Nếu không phải POST, khởi tạo form trống
        participants_data = []  # Danh sách thành viên trống
        guardians_data = []  # Danh sách phụ huynh trống

    # Tạo context cho template
    context = {
        'competition': competition,
        'form': form,
        'today': today,
        'participants_data': participants_data,
        'guardians_data': guardians_data,
    }

    return render(request, 'competition/register_team.html', context)  # Render template


@login_required
def registration_list(request, competition_id):
    # Lấy thông tin của cuộc thi cụ thể
    competition = get_object_or_404(Competition, id=competition_id)

    # Lấy tất cả các đăng ký hợp lệ cho cuộc thi này
    registrations = Registration.objects.filter(
        competition=competition,
        is_cancelled=False
    )

    # Chuyển đổi dữ liệu đăng ký thành danh sách có định dạng dễ hiển thị trong template
    registration_data = []
    for registration in registrations:
        # Kiểm tra xem đã có thông tin đội trong registration_data chưa
        team_entry = next((entry for entry in registration_data if entry['team_sbd'] == registration.team.sbd), None)

        # Nếu chưa có, tạo entry mới cho đội
        if not team_entry:
            team_entry = {
                'registration': registration,
                'team_name': registration.team.name if registration.team else 'Chưa có tên đội',
                'team_sbd': registration.team.sbd if registration.team else 'Chưa có SBD',
                'competition_group': registration.competition_group,
                'competition': competition,  # Thêm thuộc tính competition
                'team': registration.team,  # Thêm thuộc tính team
                'coach_name': registration.coach_name,  # Thêm tên huấn luyện viên
                'registration_date': registration.registration_date,  # Thêm thời gian đăng ký
                'members_participants': []
            }
            registration_data.append(team_entry)

        # Thêm thành viên vào danh sách của đội
        team_entry['members_participants'].append({
            'name': registration.student_name,
            'parent_email': registration.guardian_email
        })

    # Thực hiện phân trang
    page = request.GET.get('page', 1)  # Lấy số trang từ query parameter, mặc định là trang 1
    paginator = Paginator(registration_data, 7)  # Giới hạn mỗi trang 10 mục

    try:
        registration_data_paginated = paginator.page(page)  # Lấy trang cụ thể
    except PageNotAnInteger:
        registration_data_paginated = paginator.page(1)  # Nếu số trang không phải là số nguyên, hiển thị trang 1
    except EmptyPage:
        registration_data_paginated = paginator.page(paginator.num_pages)  # Nếu số trang quá lớn, hiển thị trang cuối cùng

    context = {
        'competition': competition,
        'registration_data': registration_data_paginated,  # Truyền dữ liệu đã phân trang
    }

    return render(request, 'competition/registration_list.html', context)


@login_required
def registration_history(request):
    user = request.user
    # Lấy tất cả các đăng ký của người dùng
    registrations = Registration.objects.filter(user=user, is_cancelled=False)

    registration_data = []
    for registration in registrations:
        # Kiểm tra xem đã có thông tin đội trong registration_data chưa
        team_entry = next((entry for entry in registration_data if entry['team_sbd'] == registration.team.sbd), None)

        # Nếu chưa có, tạo entry mới cho đội
        if not team_entry:
            team_entry = {
                'registration': registration,
                'registration_id': registration.id,  # Thêm registration_id
                'team_name': registration.team.name if registration.team else 'Chưa có tên đội',
                'team_sbd': registration.team.sbd if registration.team else 'Chưa có SBD',
                'competition_group': registration.competition_group,
                'competition': registration.competition,  # Thêm thuộc tính competition
                'team': registration.team,  # Thêm thuộc tính team
                'coach_name': registration.coach_name,  # Thêm tên huấn luyện viên
                'registration_date': registration.registration_date,  # Thêm thời gian đăng ký
                'members_participants': []  # Danh sách thành viên
            }
            registration_data.append(team_entry)

        # Thêm thành viên vào danh sách của đội
        team_entry['members_participants'].append({
            'name': registration.student_name,
            'parent_email': registration.guardian_email  # Nếu bạn cần thêm thông tin khác
        })

    # Tính số lượng bản ghi đã đăng ký thành công
    total_registrations = len(registration_data)

    # Gửi dữ liệu đến template
    return render(request, 'competition/registration_history.html', {
        'registrations': registration_data,
        'total_registrations': total_registrations
    })


@login_required
def team_detail(request, competition_id, team_id):
    # Lấy thông tin đội và các đăng ký của họ trong cuộc thi cụ thể
    registrations = Registration.objects.filter(competition_id=competition_id, team_id=team_id, is_cancelled=False)
    team_info = get_object_or_404(Team, id=team_id)  # Lấy thông tin đội thi

    context = {
        'registrations': registrations,
        'team_info': team_info  # Thêm thông tin đội vào context
    }
    return render(request, 'competition/registration_info_detail.html', context)


@login_required
def cancel_registration(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    # Lấy đội đăng ký của người dùng cho cuộc thi
    team = get_object_or_404(Team, members=request.user, competition=competition)

    # Lấy đăng ký của đội
    registration = get_object_or_404(Registration, competition=competition, team=team)

    # Đánh dấu đăng ký là đã huỷ
    registration.is_cancelled = True
    registration.save()

    # Tạo thông báo cho người dùng hiện tại khi hủy đăng ký thành công
    Notification.objects.create(
        user=request.user,
        title="Hủy đăng ký thành công",
        content=f"Bạn đã hủy đăng ký đội thi '{team.name}' cho cuộc thi {competition.name}.",
        created_at=timezone.now(),
        notification_type='user'
    )

    # Gửi thông báo cho tất cả các admin về việc hủy đăng ký
    admin_users = get_user_model().objects.filter(is_superuser=True)  # Lấy danh sách admin
    for admin in admin_users:
        Notification.objects.create(
            user=admin,
            title="Người dùng đã hủy đăng ký",
            content=f"Người dùng {request.user.username} đã hủy đăng ký đội thi '{team.name}' cho cuộc thi {competition.name}.",
            created_at=timezone.now(),
            notification_type='admin'
        )

    messages.success(request, 'Hủy đăng ký thành công!')
    return redirect('competition_detail', competition_id=competition_id)  # Quay lại trang chi tiết cuộc thi


@login_required
def add_competition(request):
    if request.method == 'POST':
        form = CompetitionForm(request.POST, request.FILES)

        # Lấy danh sách tên bảng đấu từ request POST
        group_names = request.POST.getlist('group_names[]')

        # Lấy thông tin vòng thi từ request POST
        round_names = request.POST.getlist('round_names[]')
        round_times = request.POST.getlist('round_times[]')
        round_matches = request.POST.getlist('round_matches[]')
        round_scoring_methods = request.POST.getlist('round_scoring_methods[]')

        rounds = []  # Danh sách để lưu thông tin vòng thi

        for name, time, matches, scoring_method in zip(round_names, round_times, round_matches, round_scoring_methods):
            if name:  # Kiểm tra nếu tên vòng thi không rỗng
                rounds.append({
                    'round_name': name,  # Sửa lại đây cho đúng
                    'schedule': time,  # Sửa lại đây cho đúng
                    'matches_count': int(matches) if matches else 0,  # Sửa lại đây cho đúng
                    'scoring_method': scoring_method
                })

        if form.is_valid():
            competition = form.save(commit=False)

            # Lưu danh sách nhóm
            competition.groups = [{"group_name": name} for name in group_names]
            # Lưu thông tin vòng thi dưới dạng JSON
            competition.rounds = rounds  # Nếu bạn có trường `rounds` trong model Competition
            competition.save()

            # Tạo thông báo cho người dùng
            CustomUser = get_user_model()
            users = CustomUser.objects.filter(is_superuser=False)
            for user in users:
                Notification.objects.create(
                    user=user,
                    title="Cuộc thi mới",
                    content=f"Thông tin về cuộc thi {competition.name} đã chính thức xuất hiện. Hãy tìm hiểu và tham gia ngay!",
                    created_at=timezone.now(),
                    notification_type='user'
                )

            admin_users = CustomUser.objects.filter(is_superuser=True)
            for admin in admin_users:
                Notification.objects.create(
                    user=admin,
                    title=f"Cuộc thi mới {competition.name}",
                    content=f"Cuộc thi {competition.name} đã được {request.user.get_full_name()} tạo.",
                    created_at=timezone.now(),
                    notification_type='admin'
                )

            messages.success(request, f"Cuộc thi '{competition.name}' đã được tạo thành công!")
            return redirect('contest_list')
        else:
            print(form.errors)

    else:
        form = CompetitionForm()

    return render(request, 'competition/add_competition.html', {
        'form': form,
    })


@login_required
def edit_competition(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    if request.method == 'POST':
        form = CompetitionForm(request.POST, request.FILES, instance=competition)

        # Lấy danh sách tên bảng đấu từ request POST
        group_names = request.POST.getlist('group_names[]')

        # Lấy thông tin vòng thi từ request POST
        round_names = request.POST.getlist('round_names[]')
        round_times = request.POST.getlist('round_times[]')
        round_matches = request.POST.getlist('round_matches[]')
        round_scoring_methods = request.POST.getlist('round_scoring_methods[]')

        rounds = []  # Danh sách để lưu thông tin vòng thi

        for name, time, matches, scoring_method in zip(round_names, round_times, round_matches, round_scoring_methods):
            if name:  # Kiểm tra nếu tên vòng thi không rỗng
                rounds.append({
                    'round_name': name,
                    'schedule': time,
                    'matches_count': int(matches) if matches else 0,
                    'scoring_method': scoring_method
                })

        if form.is_valid():
            competition = form.save(commit=False)

            # Lưu danh sách nhóm
            competition.groups = [{"group_name": name} for name in group_names]
            # Lưu thông tin vòng thi dưới dạng JSON
            competition.rounds = rounds

            competition.save()

            messages.success(request, f"Cuộc thi '{competition.name}' đã được cập nhật thành công!")
            return redirect('contest_list')
        else:
            print(form.errors)

    else:
        form = CompetitionForm(instance=competition)

    return render(request, 'competition/edit_competition.html', {
        'form': form,
        'competition': competition,  # Truyền thông tin cuộc thi cho template
    })


@login_required
def delete_competition(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    if request.method == 'POST':
        competition_name = competition.name  # Lưu lại tên cuộc thi trước khi xóa
        competition.delete()

        # Tạo thông báo cho tất cả người dùng thường
        users = CustomUser.objects.filter(is_superuser=False)  # Lọc người dùng thường (không phải admin)
        for user in users:
            Notification.objects.create(
                user=user,
                title="Cuộc thi đã bị xóa",
                content=f"Cuộc thi {competition_name} đã bị xóa. Hãy theo dõi các cuộc thi khác!",
                created_at=timezone.now(),
                notification_type='user'
            )

        # Tạo thông báo cho admin về việc cuộc thi đã bị xóa
        admin_users = CustomUser.objects.filter(is_superuser=True)  # Lấy danh sách admin
        for admin in admin_users:
            Notification.objects.create(
                user=admin,
                title="Cuộc thi đã bị xóa",
                content=f"Người dùng {request.user.get_full_name()} đã xóa cuộc thi {competition_name}.",
                created_at=timezone.now(),
                notification_type='admin'
            )

        messages.success(request, f"Cuộc thi '{competition_name}' đã được xóa thành công.")
        return redirect('contest_list')

    return render(request, 'competition/delete_competition.html', {'competition': competition})


@login_required
def add_result(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)
    rounds_data = competition.rounds or []
    groups_data = competition.groups or []

    selected_round = request.GET.get('round', rounds_data[0]['round_name'] if rounds_data else None)
    selected_group = request.GET.get('group', groups_data[0]['group_name'] if groups_data else None)

    matches_count = 0
    scoring_method = "Tổng điểm toàn bộ các lượt thi"

    if selected_round and selected_group:
        selected_round_data = next((r for r in rounds_data if r['round_name'] == selected_round), None)
        matches_count = selected_round_data.get('matches_count', 0) if selected_round_data else 0
        scoring_method = selected_round_data.get('scoring_method', scoring_method)

    # Lấy danh sách các đội trong bảng đấu
    teams = Team.objects.filter(
        id__in=Registration.objects.filter(
            competition=competition,
            competition_group=selected_group
        ).values('team_id')
    )

    # Lấy thông tin thành viên của các đội
    for team in teams:
        registrations = Registration.objects.filter(team=team)
        team.members_list = [
            {
                'name': reg.student_name,
                'parent_email': reg.guardian_email
            }
            for reg in registrations
        ]
        print(f"Team {team.name} Members List:", team.members_list)

    if request.method == 'POST':
        team_ids = request.POST.getlist('team_ids')
        if not team_ids:
            messages.error(request, "Không có đội nào được gửi để lưu kết quả.")
            return redirect(request.path)

        for team_id in team_ids:
            team_score = []
            for i in range(1, matches_count + 1):
                score_field = f'score_{team_id}_{i}'  # Trường lưu điểm của đội
                score = request.POST.get(score_field, '')  # Đặt giá trị mặc định là chuỗi rỗng
                print(f"Score for Team {team_id}, Match {i}:", score)

                # Kiểm tra nếu score không phải là chuỗi rỗng và chuyển thành float
                if score.strip():  # Nếu có giá trị hợp lệ
                    team_score.append(float(score))
                else:  # Nếu chuỗi rỗng hoặc không hợp lệ
                    team_score.append(0)

            print(f"Team {team_id} Scores:", team_score)

            ranking_score = sum(team_score) if scoring_method == "Tổng điểm toàn bộ các lượt thi" else max(team_score)

            try:
                existing_result = CompetitionResult.objects.filter(
                    team_id=team_id, competition=competition,
                    round_name=selected_round, group_name=selected_group
                ).first()

                if existing_result:
                    existing_result.score = team_score
                    existing_result.ranking_score = ranking_score
                    existing_result.save()
                else:
                    CompetitionResult.objects.create(
                        competition=competition,
                        team_id=team_id,
                        round_name=selected_round,
                        group_name=selected_group,
                        score=team_score,
                        ranking_score=ranking_score
                    )
            except Exception as e:
                messages.error(request, f"Lỗi khi lưu kết quả cho đội {team_id}: {e}")
                continue

        messages.success(request, "Kết quả đã được lưu thành công.")
        return redirect('result_detail', competition_id=competition.id)

    return render(request, 'result/add_result.html', {
        'competition': competition,
        'rounds_data': rounds_data,
        'groups_data': groups_data,
        'teams': teams,
        'selected_round': selected_round,
        'selected_group': selected_group,
        'matches_count': matches_count,
    })


@login_required
def result_list(request):
    competitions = Competition.objects.all()  # Lấy tất cả các cuộc thi
    return render(request, 'result/result_list.html', {'competitions': competitions})


@login_required
def result_detail(request, competition_id):
    # Lấy đối tượng cuộc thi theo ID
    competition = get_object_or_404(Competition, id=competition_id)

    # Lấy thông tin bảng đấu và vòng đấu từ cuộc thi
    groups = competition.groups  # Ví dụ: [{"group_name": "Bảng Tiểu học"}, {"group_name": "Bảng THCS"}]
    rounds = competition.rounds  # Ví dụ: [{"round_name": "Vòng 1"}, {"round_name": "Vòng 2"}]

    # Tổ chức dữ liệu kết quả theo cấu trúc {group_name: {round_name: [kết quả]}}
    results = {}
    for group in groups:
        group_name = group.get('group_name')  # Lấy tên bảng đấu
        results[group_name] = {}

        for round_info in rounds:
            round_name = round_info.get('round_name')  # Lấy tên vòng đấu
            # Truy vấn kết quả của các đội thi cho bảng và vòng hiện tại, sắp xếp theo điểm giảm dần
            round_results = CompetitionResult.objects.filter(
                competition=competition,
                group_name=group_name,
                round_name=round_name
            ).order_by('-score')

            # Lưu kết quả vào cấu trúc dữ liệu
            results[group_name][round_name] = []

            # Lấy thông tin thành viên cho từng kết quả
            for result in round_results:
                # Lấy thông tin đăng ký cho đội thi
                registrations = Registration.objects.filter(team=result.team)

                # Tạo danh sách thành viên của đội
                team_members = []
                for registration in registrations:
                    team_members.append({
                        'name': registration.student_name,
                        'parent_email': registration.guardian_email
                    })

                # Lưu thông tin kết quả và thành viên
                results[group_name][round_name].append({
                    'result': result,
                    'members': team_members
                })

    # Truyền dữ liệu đến template
    return render(request, 'result/result_detail.html', {
        'competition': competition,
        'results': results,
    })


@login_required
def edit_result(request, competition_id, result_id):
    # Lấy đối tượng Competition và CompetitionResult
    competition = get_object_or_404(Competition, id=competition_id)
    result = get_object_or_404(CompetitionResult, id=result_id, competition=competition)

    # Dữ liệu vòng và bảng từ Competition (JSONField)
    rounds_data = competition.rounds or []
    groups_data = competition.groups or []

    # Lấy vòng đấu, bảng đấu và thông tin số trận
    selected_round = result.round_name
    selected_group = result.group_name

    selected_round_data = next((r for r in rounds_data if r['round_name'] == selected_round), None)
    matches_count = selected_round_data.get('matches_count', 0) if selected_round_data else 0
    scoring_method = selected_round_data.get('scoring_method', 'Tổng điểm toàn bộ các lượt thi')

    # Lấy danh sách các đội trong bảng đấu
    teams = Team.objects.filter(
        id__in=Registration.objects.filter(
            competition=competition,
            competition_group=selected_group
        ).values('team_id')
    )

    # Chuẩn bị dữ liệu điểm số cho từng đội
    for team in teams:
        registrations = Registration.objects.filter(team=team)
        team_members = []

        for registration in registrations:
            team_members.append({
                'name': registration.student_name,
                # Nếu cần email phụ huynh, bạn có thể bỏ comment ở đây
                # 'parent_email': registration.guardian_email
            })

        team.members_list = team_members  # Gán danh sách thành viên vào team

        result_data = CompetitionResult.objects.filter(
            team_id=team.id,
            competition=competition,
            round_name=selected_round,
            group_name=selected_group
        ).first()

        # Nếu có điểm số, đảm bảo đúng chiều dài, ngược lại tạo danh sách mặc định
        if result_data and isinstance(result_data.score, list):
            team.scores = result_data.score + [0] * (matches_count - len(result_data.score))
        else:
            team.scores = [0] * matches_count

    # Xử lý POST request để cập nhật kết quả
    if request.method == 'POST':
        for team in teams:
            team_score = []
            for i in range(matches_count):
                score_field = f'score_{team.id}_{i+1}'
                score = request.POST.get(score_field, '').strip()

                # Nếu có giá trị hợp lệ, chuyển thành float, ngược lại gán 0
                team_score.append(float(score) if score else 0)

            # Tính điểm ranking_score
            ranking_score = (
                sum(team_score) if scoring_method == "Tổng điểm toàn bộ các lượt thi"
                else max(team_score, default=0)
            )

            print(f"Updated Scores for Team {team.id}:", team_score)  # Debug điểm số mới
            print(f"Updated Ranking Score for Team {team.id}:", ranking_score)  # Debug điểm ranking mới

            # Cập nhật hoặc tạo mới đối tượng CompetitionResult
            result_data, created = CompetitionResult.objects.get_or_create(
                team_id=team.id,
                competition=competition,
                round_name=selected_round,
                group_name=selected_group,
                defaults={'score': team_score, 'ranking_score': ranking_score}
            )

            if not created:
                result_data.score = team_score
                result_data.ranking_score = ranking_score
                result_data.save()

        messages.success(request, "Kết quả đã được cập nhật thành công.")
        return redirect('result_detail', competition_id=competition.id)

    # Truyền thông tin vào template
    return render(request, 'result/edit_result.html', {
        'competition': competition,
        'rounds_data': rounds_data,
        'groups_data': groups_data,
        'teams': teams,
        'selected_round': selected_round,
        'selected_group': selected_group,
        'matches_count': matches_count,
        'result': result,
    })


@login_required
def kit_list(request):
    kits = Kit.objects.all()
    return render(request, 'kit/kit_list.html', {'kits': kits})


@login_required
def kit_create(request):
    if request.method == 'POST':
        form = KitForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bộ thiết bị đã được tạo thành công.')
            return redirect('kit_list')
    else:
        form = KitForm()
    return render(request, 'kit/kit_form.html', {'form': form})


@login_required
def kit_update(request, pk):
    kit = get_object_or_404(Kit, pk=pk)  # Sử dụng get_object_or_404 để đảm bảo an toàn
    if request.method == 'POST':
        form = KitForm(request.POST, request.FILES, instance=kit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bộ thiết bị đã được cập nhật thành công.')
            return redirect('kit_list')
    else:
        form = KitForm(instance=kit)  # Đảm bảo giữ giá trị hiện tại

    return render(request, 'kit/kit_form.html', {'form': form, 'kit': kit})


@login_required
def kit_delete(request, pk):
    kit = Kit.objects.get(pk=pk)
    if request.method == 'POST':
        kit.delete()
        messages.success(request, 'Bộ thiết bị đã bị xóa thành công.')
        return redirect('kit_list')
    return render(request, 'kit/delete_kit.html', {'kit': kit})


@login_required
def kit_detail(request, pk):
    kit = get_object_or_404(Kit, pk=pk)
    images = kit.images.all()  # Lấy tất cả hình ảnh liên quan đến bộ kit

    if request.method == 'POST':
        image_form = KitImageForm(request.POST, request.FILES)
        if image_form.is_valid():
            new_image = image_form.save(commit=False)
            new_image.kit = kit  # Gán hình ảnh vào bộ kit
            new_image.save()
            return redirect('kit_detail', pk=pk)  # Reload lại trang sau khi upload
    else:
        image_form = KitImageForm()

    return render(request, 'kit/kit_detail.html', {'kit': kit, 'images': images, 'image_form': image_form})


@login_required
def add_image(request, kit_id):
    kit = get_object_or_404(Kit, id=kit_id)

    if request.method == 'POST':
        form = KitImageForm(request.POST, request.FILES)
        if form.is_valid():
            kit_image = form.save(commit=False)
            kit_image.kit = kit
            kit_image.save()
            return redirect('kit_detail', pk=kit.id)  # Chuyển hướng sau khi upload thành công
    else:
        form = KitImageForm()

    return render(request, 'kit/add_image.html', {'form': form, 'kit': kit})


@login_required
def sponsor_list(request):
    sponsors = Sponsor.objects.all()
    return render(request, 'sponsor/sponsor_list.html', {'sponsors': sponsors})


@login_required
def add_sponsor(request):
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('sponsor_list')
    else:
        form = SponsorForm()
    return render(request, 'sponsor/add_sponsor.html', {'form': form})


@login_required
def edit_sponsor(request, sponsor_id):
    sponsor = get_object_or_404(Sponsor, id=sponsor_id)
    if request.method == 'POST':
        form = SponsorForm(request.POST, request.FILES, instance=sponsor)
        if form.is_valid():
            form.save()
            return redirect('sponsor_list')  # Redirect đến danh sách nhà tài trợ
    else:
        form = SponsorForm(instance=sponsor)
    return render(request, 'sponsor/edit_sponsor.html', {'form': form})


@login_required
def delete_sponsor(request, sponsor_id):
    sponsor = get_object_or_404(Sponsor, id=sponsor_id)
    if request.method == 'POST':
        sponsor.delete()
        return redirect('sponsor_list')  # Redirect đến danh sách nhà tài trợ
    return render(request, 'sponsor/delete_sponsor.html', {'sponsor': sponsor})


@login_required
def submit_feedback(request):
    if request.method == "POST":
        form = FeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.user = request.user

            # Xử lý ảnh
            if feedback.image:
                try:
                    img = Image.open(feedback.image)
                    img = img.convert('RGB')  # Chuyển đổi sang RGB

                    # Lưu vào một buffer
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='JPEG')  # Lưu ảnh dưới dạng JPEG
                    img_byte_arr.seek(0)

                    # Tạo một ContentFile từ buffer
                    feedback.image.save(f"{feedback.image.name}.jpg", ContentFile(img_byte_arr.read()), save=False)
                except Exception as e:
                    form.add_error('image', 'Lỗi khi xử lý hình ảnh: ' + str(e))
                    messages.error(request, "Có lỗi xảy ra khi xử lý hình ảnh.")
                    return render(request, 'feedback/submit_feedback.html', {'form': form})

            feedback.save()
            messages.success(request, "Phản hồi của bạn đã được gửi thành công!")
            return redirect('home')  # Chuyển hướng về trang home sau khi gửi phản hồi
        else:
            # In ra lỗi để dễ debug
            print(form.errors)  # In ra các lỗi của form trong console
            messages.error(request, "Có lỗi xảy ra khi gửi phản hồi. Vui lòng kiểm tra lại.")
    else:
        form = FeedbackForm()

    return render(request, 'feedback/submit_feedback.html', {'form': form})


@login_required
def feedback_list(request):
    if request.user.is_superuser:
        feedback_list = Feedback.objects.all().order_by('-created_at')
        return render(request, 'feedback/feedback_list.html', {'feedback_list': feedback_list})
    else:
        messages.error(request, "Bạn không có quyền truy cập vào trang này.")
        return redirect('home')


@login_required
def export_feedback_to_excel(request):
    # Lấy danh sách phản hồi (feedback)
    feedback_list = Feedback.objects.all().order_by('-created_at')

    # Tạo workbook và worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Report"

    # Đặt font mặc định cho toàn bộ bảng (size 14)
    default_font = Font(size=14)

    # Đặt tiêu đề cột với font bôi đậm và kích thước 14, không có màu nền
    columns = ['STT', 'Chủ đề', 'Nội dung phản hồi', 'Đánh giá', 'Hình ảnh', 'Ngày phản hồi', 'Trạng thái']

    # Đảm bảo bôi đậm tiêu đề cột với font đậm và không có màu nền
    for col_num, column_title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True, size=14, color='000000')  # Đặt font đậm, size 14, không có màu nền
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Căn giữa tiêu đề

    # Thêm dữ liệu vào bảng từ dòng thứ hai
    for index, feedback in enumerate(feedback_list, start=1):
        created_at_naive = make_naive(feedback.created_at)  # Chuyển timezone-aware datetime thành timezone-naive
        row = [
            index,
            feedback.subject,
            feedback.content.replace('\r\n', '\n'),  # Thay thế xuống dòng
            feedback.rating,
            '',
            created_at_naive.strftime("%d/%m/%Y"),
            'Đã xem' if feedback.is_viewed else 'Chưa xem'
        ]
        ws.append(row)

        # Đảm bảo nội dung phản hồi được wrap text
        ws.cell(row=index + 1, column=3).alignment = Alignment(wrap_text=True)  # Cột "Nội dung phản hồi"

        # Nếu có hình ảnh, thêm vào ô tương ứng
        if feedback.image:
            img_path = feedback.image.path
            if os.path.exists(img_path):
                img = OpenPyXLImage(img_path)

                # Điều chỉnh kích thước ảnh
                img.width = 300  # Xác định chiều rộng ảnh
                img.height = 200  # Xác định chiều cao ảnh

                # Thêm hình ảnh vào ô (cột 5 là "Hình ảnh")
                ws.add_image(img, f'E{index + 1}')

                # Cố định độ cao hàng dựa trên chiều cao của ảnh (1 đơn vị = 0.75 pixel)
                row_height = img.height * 0.75  # Chuyển đổi chiều cao ảnh sang đơn vị hàng
                ws.row_dimensions[index + 1].height = row_height  # Điều chỉnh độ cao hàng

    # Đặt độ rộng fix cứng cho cột "Hình ảnh"
    ws.column_dimensions['E'].width = 45  # 45 tương đương với khoảng 270 pixels

    # Điều chỉnh độ rộng các cột khác dựa trên dữ liệu
    for col_num, column_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_num)

        # Lấy độ dài lớn nhất giữa tiêu đề cột và dữ liệu trong cột đó
        max_length = max([len(str(cell.value)) if cell.value else 0 for cell in column_cells])

        # Đặt độ rộng cho cột, cộng thêm 2 để dễ nhìn
        if col_letter != 'E':  # Trừ cột E vì đã fix cứng độ rộng
            ws.column_dimensions[col_letter].width = max_length + 2

        # Cập nhật kích thước chữ và căn giữa cho từng ô
        for cell in column_cells:
            cell.font = default_font  # Tăng kích thước chữ lên 14 cho toàn bộ dữ liệu
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Căn giữa toàn bộ dữ liệu

    # Tạo một BytesIO để lưu workbook
    output = BytesIO()
    wb.save(output)  # Ghi workbook vào BytesIO
    output.seek(0)  # Đặt con trỏ về đầu để có thể đọc từ đầu

    # Tạo tên file theo ngày tháng
    filename = f"feedback_report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    # Thiết lập phản hồi HTTP để trả file Excel về cho người dùng
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Reset lại dữ liệu phản hồi sau khi xuất ra file
    feedback_list.delete()

    return response


@login_required
def toggle_viewed_status(request, feedback_id):
    feedback = get_object_or_404(Feedback, id=feedback_id)

    if request.method == 'POST':
        # Đảo ngược giá trị của is_viewed
        feedback.is_viewed = not feedback.is_viewed
        feedback.save()

        # Trả về JSON với trạng thái mới
        return JsonResponse({'is_viewed': feedback.is_viewed})

    # Nếu không phải POST, trả về lỗi không hợp lệ
    return JsonResponse({'error': 'Invalid request'}, status=400)


def sanitize_filename(filename):
    # Bước 1: Loại bỏ các từ không cần thiết như "Cuộc thi"
    filename = filename.replace("Cuộc thi", "").strip()  # Loại bỏ "Cuộc thi" và trim khoảng trắng

    # Bước 2: Thay thế các ký tự không hợp lệ trong tên file (chỉ giữ lại các ký tự an toàn)
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)  # Thay thế các ký tự không hợp lệ bằng dấu gạch dưới
    filename = filename.replace(' ', '_')  # Thay thế dấu cách bằng dấu gạch dưới

    # Bước 3: Trả về tên file đã xử lý
    return filename


@login_required
def export_registrations_to_excel(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    # Lấy danh sách các bảng thi từ cuộc thi
    competition_groups = Registration.objects.filter(competition=competition).values_list('competition_group', flat=True).distinct()

    # Tạo workbook
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    # Kiểm tra nếu không có bảng thi nào, tạo một sheet mặc định
    if not competition_groups:
        ws = wb.create_sheet(title="Không có dữ liệu")
        ws.append(['Không có dữ liệu đăng ký'])

    # Duyệt qua các bảng thi (competition_group) và tạo các sheet riêng
    for group in competition_groups:
        group_name = group if group else 'Chưa phân bảng'
        sheet_title = f'{group_name}'[:31]  # Giới hạn tên sheet tối đa 31 ký tự
        ws = wb.create_sheet(title=sheet_title)

        # Thêm tiêu đề cột vào sheet
        columns = ['STT', 'Tên đội', 'Số báo danh', 'Huấn luyện viên', 'Thí sinh', 'Thời gian đăng ký']
        ws.append(columns)

        # Bôi đậm và căn giữa tiêu đề cột
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Lọc các đăng ký cho bảng thi hiện tại
        group_registrations = Registration.objects.filter(competition=competition, competition_group=group)

        # Duyệt qua các đội thi và thêm dữ liệu vào sheet
        seen_teams = set()  # Sử dụng set để đảm bảo mỗi đội chỉ xuất hiện một lần
        for idx, registration in enumerate(group_registrations, start=1):
            team = registration.team
            if team not in seen_teams:
                # Đánh dấu đội đã được xử lý
                seen_teams.add(team)

                # Tạo danh sách thí sinh cho đội (gộp các thành viên nếu có)
                team_members = ", ".join([reg.student_name for reg in group_registrations if reg.team == team])

                # Chuyển thời gian đăng ký sang múi giờ địa phương
                local_registration_time = localtime(registration.registration_date).strftime('%H:%M, %d/%m/%Y')

                # Thêm hàng dữ liệu cho đội
                row = [
                    idx,
                    team.name if team.name else 'Chưa có tên đội',
                    team.sbd if team.sbd else 'Chưa có SBD',
                    registration.coach_name if registration.coach_name else 'Chưa có huấn luyện viên',
                    team_members,  # Thí sinh
                    local_registration_time  # Thời gian đăng ký
                ]
                ws.append(row)

        # Căn giữa nội dung các ô
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Điều chỉnh độ rộng của cột cho từng sheet
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Lưu workbook vào file ảo
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Xử lý tên file
    safe_filename = sanitize_filename(competition.name)

    # Tạo tên file hoàn chỉnh
    filename = f'Danh_sach_dang_ky_{safe_filename}.xlsx'

    # Thiết lập response để tải về file Excel
    response = HttpResponse(file_stream,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response


@login_required
def export_results_to_excel(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    # Lấy thông tin bảng đấu và vòng đấu từ cuộc thi
    groups = competition.groups  # [{"group_name": "Bảng Tiểu học"}, {"group_name": "Bảng THCS"}]
    rounds = competition.rounds  # [{"round_name": "Vòng loại"}, {"round_name": "Vòng chung kết"}]

    # Tạo workbook
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    # Duyệt qua từng bảng đấu và vòng đấu để tạo các sheet riêng
    for group in groups:
        group_name = group.get('group_name')
        for round_info in rounds:
            round_name = round_info.get('round_name')

            # Tạo tên sheet dựa trên bảng và vòng đấu
            sheet_title = f'{group_name}_{round_name}'[:31]  # Giới hạn tên sheet tối đa 31 ký tự
            ws = wb.create_sheet(title=sheet_title)

            # Thêm tiêu đề cột cho sheet
            columns = ['Thứ hạng', 'Số báo danh', 'Tên đội', 'Huấn luyện viên', 'Thành viên', 'Điểm xếp hạng']
            ws.append(columns)

            # Bôi đậm và căn giữa tiêu đề cột
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Truy vấn kết quả của các đội thi cho bảng và vòng hiện tại
            round_results = CompetitionResult.objects.filter(
                competition=competition,
                group_name=group_name,
                round_name=round_name
            ).order_by('-score')

            # Duyệt qua các kết quả để thêm vào sheet
            for rank, result in enumerate(round_results, start=1):
                # Lấy thông tin đăng ký của đội
                registrations = Registration.objects.filter(team=result.team)

                # Tạo danh sách tên thành viên
                team_members = ", ".join([reg.student_name for reg in registrations])

                # Thêm hàng dữ liệu
                row = [
                    rank,
                    result.team.sbd,  # Số báo danh
                    result.team.name,  # Tên đội
                    result.team.coach,  # Huấn luyện viên
                    team_members,  # Thành viên
                    result.ranking_score  # Điểm xếp hạng
                ]
                ws.append(row)

            # Căn giữa nội dung các ô
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Điều chỉnh độ rộng của cột cho từng sheet
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Lưu workbook vào file ảo
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Thiết lập response để tải về file Excel
    safe_filename = sanitize_filename(competition.name)  # Bạn có thể viết hàm sanitize_filename nếu cần
    response = HttpResponse(file_stream,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Ket_qua_{safe_filename}.xlsx'

    return response


@login_required
def competition_guide(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)

    if request.method == 'POST':
        guide_file = request.FILES.get('guide_file')
        document_name = request.POST.get('document_name')
        note = request.POST.get('note')
        if guide_file:
            guide_file_instance = GuideFile(
                competition=competition,
                file=guide_file,
                original_file_name=guide_file.name,  # Lưu tên gốc vào original_file_name
                document_name=document_name,
                note=note,
                is_confirmed=False
            )
            guide_file_instance.save()
            return redirect('confirm_guide_file', guide_file_id=guide_file_instance.id)

    # Lấy danh sách tài liệu đã được xác nhận
    guide_files = competition.guide_files.filter(is_confirmed=True).order_by('-uploaded_at')

    # Thêm thông tin về xem trước vào context
    for guide_file in guide_files:
        # Lấy loại MIME của tệp tin
        mime_type, encoding = mimetypes.guess_type(guide_file.file.name)

        # Nếu là ảnh hoặc PDF, thêm thông tin vào file
        guide_file.is_previewable = False
        if mime_type:
            if mime_type.startswith('image/') or mime_type == 'application/pdf':
                guide_file.is_previewable = True

    return render(request, 'competition/competition_guide.html', {
        'competition': competition,
        'guide_files': guide_files
    })


@login_required
def confirm_guide_file(request, guide_file_id):
    guide_file = get_object_or_404(GuideFile, id=guide_file_id)

    if request.method == 'POST':
        guide_file.is_confirmed = True  # Đánh dấu là đã xác nhận
        guide_file.save()
        return redirect('competition_guide', competition_id=guide_file.competition.id)  # Quay lại trang danh sách

    return render(request, 'competition/confirm_guide_file.html', {'guide_file': guide_file})


@login_required
def edit_guide_file(request, pk):
    guide_file = get_object_or_404(GuideFile, pk=pk)

    if request.method == 'POST':
        form = GuideFileForm(request.POST, request.FILES, instance=guide_file)

        if form.is_valid():
            guide_file.note = form.cleaned_data['note']

            # Kiểm tra xem có tệp mới không
            if 'file' in request.FILES and request.FILES['file']:
                guide_file.file = request.FILES['file']  # Cập nhật tệp mới
                guide_file.original_file_name = request.FILES['file'].name  # Cập nhật tên gốc

            guide_file.save()  # Lưu thay đổi
            return redirect('competition_guide', competition_id=guide_file.competition.id)
    else:
        form = GuideFileForm(instance=guide_file)

    # Sử dụng `original_file_name` để hiển thị tên gốc của tệp
    current_file_name = guide_file.original_file_name if guide_file.original_file_name else "Chưa có tệp"

    return render(request, 'competition/edit_guide_file.html', {
        'form': form,
        'guide_file': guide_file,
        'current_file_name': current_file_name
    })


@login_required
def delete_guide_file(request, pk):
    guide_file = get_object_or_404(GuideFile, pk=pk)
    competition_id = guide_file.competition.id
    guide_file.delete()
    return redirect('competition_guide', competition_id=competition_id)


@login_required
def download_guide_file(request, guide_file_id):
    guide_file = get_object_or_404(GuideFile, id=guide_file_id)
    response = FileResponse(guide_file.file.open('rb'), as_attachment=True)
    response['Content-Disposition'] = f'attachment; filename="{guide_file.original_file_name}"'
    return response


User = get_user_model()


@login_required
def manage_users(request):
    # Kiểm tra xem người dùng có phải admin không
    if request.user.role != 'admin':
        return HttpResponseForbidden("Bạn không có quyền truy cập vào trang này.")

    # Lấy dữ liệu tìm kiếm từ form
    search_query = request.GET.get('search', '')

    # Tạo truy vấn tìm kiếm
    if search_query:
        users = User.objects.filter(
            Q(username__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(role__icontains=search_query)
        )
    else:
        users = User.objects.all()

    # Sắp xếp dữ liệu
    users = users.order_by('username')

    # Phân trang: mỗi trang sẽ hiển thị 8 người dùng
    paginator = Paginator(users, 8)  # 8 là số người dùng mỗi trang
    page_number = request.GET.get('page')  # Lấy số trang từ query string

    # Kiểm tra xem page_number có hợp lệ không
    try:
        page_obj = paginator.get_page(page_number)
    except ValueError:
        page_obj = paginator.get_page(1)  # Nếu có lỗi, dùng trang 1

    # Xử lý việc thay đổi vai trò người dùng nếu có
    if request.method == 'POST':
        username = request.POST.get('username')
        role = request.POST.get('role')
        user = get_object_or_404(User, username=username)
        user.role = role
        user.save()

    # Trả lại template với đối tượng phân trang
    return render(request, 'manage_users.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })


@login_required
def edit_registration(request, registration_id):
    # Lấy bản đăng ký học sinh cần chỉnh sửa theo registration_id
    registration = get_object_or_404(Registration, id=registration_id)

    # Kiểm tra quyền truy cập
    if registration.user != request.user:
        messages.error(request, 'Bạn không có quyền chỉnh sửa thông tin này.')
        return redirect('competition_detail')

    # Lấy tất cả các bản ghi của đội thi này theo `team_id`
    registrations = Registration.objects.filter(
        competition_id=registration.competition.id,
        team_id=registration.team.id,
        user_id=request.user.id
    )

    if request.method == 'POST':
        # Cập nhật các thông tin chung của đội thi (áp dụng cho toàn bộ bản ghi của đội)
        region = request.POST.get('region', registration.region)
        city = request.POST.get('city', registration.city)
        team_name = request.POST.get('team_name', registration.team_name)
        team_email = request.POST.get('team_email', registration.team_email)
        competition_group = request.POST.get('competition_group', registration.competition_group)
        coach_name = request.POST.get('coach_name', registration.coach_name)
        coach_unit = request.POST.get('coach_unit', registration.coach_unit)
        coach_phone = request.POST.get('coach_phone', registration.coach_phone)

        # Lặp qua tất cả các bản ghi của đội và cập nhật thông tin chung
        for reg in registrations:
            reg.region = region
            reg.city = city
            reg.team_name = team_name
            reg.team_email = team_email
            reg.competition_group = competition_group
            reg.coach_name = coach_name
            reg.coach_unit = coach_unit
            reg.coach_phone = coach_phone
            reg.save()  # Lưu cập nhật chung cho mỗi bản ghi của đội

        # Cập nhật thông tin học sinh và phụ huynh riêng biệt cho từng `registration`
        for i, reg in enumerate(registrations, start=1):
            reg.guardian_name = request.POST.get(f'guardian_name_{i}', reg.guardian_name)
            reg.relationship = request.POST.get(f'relationship_{i}', reg.relationship)
            reg.guardian_phone = request.POST.get(f'guardian_phone_{i}', reg.guardian_phone)
            reg.guardian_email = request.POST.get(f'guardian_email_{i}', reg.guardian_email)

            reg.student_name = request.POST.get(f'student_name_{i}', reg.student_name)
            reg.student_phone = request.POST.get(f'student_phone_{i}', reg.student_phone)
            reg.student_email = request.POST.get(f'student_email_{i}', reg.student_email)
            reg.student_class = request.POST.get(f'student_class_{i}', reg.student_class)
            reg.birth_year = int(request.POST.get(f'birth_year_{i}', reg.birth_year))
            reg.gender = request.POST.get(f'gender_{i}', reg.gender)
            reg.ethnicity = request.POST.get(f'ethnicity_{i}', reg.ethnicity)
            reg.address = request.POST.get(f'address_{i}', reg.address)

            # Cập nhật ảnh nếu có ảnh mới được tải lên
            student_photo = request.FILES.get(f'student_photo_{i}', None)
            if student_photo:
                reg.student_photo = student_photo

            reg.save()  # Lưu từng bản ghi học sinh sau khi cập nhật thông tin

        messages.success(request, 'Thông tin đội thi đã được cập nhật thành công.')
        return redirect('registration_history')

    else:
        # Trả về template với thông tin đã lọc
        context = {
            'registration': registration,
            'range_list': range(1, registration.member_count + 1),
            'registrations': registrations,
        }
        return render(request, 'competition/edit_registration.html', context)


def password_reset_request(request):
    if request.method == "POST":
        username = request.POST.get("username")  # Lấy username từ form
        email = request.POST.get("email")  # Lấy email từ form
        phone_number = request.POST.get("phone_number")  # Lấy số điện thoại từ form

        try:
            # Tìm kiếm người dùng với username, email và phone_number
            user = CustomUser.objects.get(username=username, email=email, phone_number=phone_number)
            print(f"DEBUG: Tìm thấy user: Username={username}, Email={user.email}, Phone={user.phone_number}")

            # Tạo mã code ngẫu nhiên cho người dùng
            code = generate_random_code()

            # Cập nhật mật khẩu tạm thời của người dùng thành mã code này
            user.set_password(code)
            user.save()

            # Gửi email với mã code
            subject = "Mã reset mật khẩu của bạn"
            email_template_name = "account/password_reset_email.html"
            context = {
                "user": user,
                "code": code,
            }

            email_message = render_to_string(email_template_name, context)

            # Gửi email với nội dung HTML
            send_mail(
                subject,
                'Forget password',
                'minhtcppdev@gmail.com',  # Địa chỉ email gửi đi
                [user.email],
                html_message=email_message,
                fail_silently=False
            )
            # print(f"DEBUG: Email reset mật khẩu đã được gửi đến {user.email}")
            # messages.success(request, "Mã reset mật khẩu đã được gửi vào email của bạn.")
            return redirect("login")

        except CustomUser.DoesNotExist:
            # print(f"DEBUG: Không tìm thấy user với Username={username}, Email={email}, Phone={phone_number}")
            messages.error(request, "Thông tin không chính xác. Vui lòng kiểm tra lại.")

    return render(request, "account/password_reset_form.html")


# def password_reset_confirm(request):
#     if request.method == "POST":
#         code = request.POST.get("code")
#         new_password1 = request.POST.get("new_password1")
#         new_password2 = request.POST.get("new_password2")
#
#         # Kiểm tra mã code hợp lệ
#         try:
#             user = CustomUser.objects.get(password=code)  # Kiểm tra xem mật khẩu tạm thời có khớp không
#             if new_password1 == new_password2:
#                 # Cập nhật mật khẩu người dùng
#                 user.set_password(new_password1)
#                 user.save()
#                 print(f"DEBUG: Mật khẩu mới được lưu cho user: {user.username}")
#
#                 # Đăng nhập người dùng ngay sau khi thay đổi mật khẩu
#                 login(request, user)
#                 messages.success(request, "Mật khẩu đã được đặt lại thành công!")
#                 return redirect("password_reset_complete")
#             else:
#                 messages.error(request, "Mật khẩu không khớp. Vui lòng thử lại.")
#         except CustomUser.DoesNotExist:
#             messages.error(request, "Mã reset mật khẩu không hợp lệ.")
#
#     return render(request, "account/password_reset_confirm.html")


# def password_reset_done(request):
#     return render(request, "account/password_reset_done.html")
#
#
# def password_reset_complete(request):
#     return render(request, "account/password_reset_complete.html")

